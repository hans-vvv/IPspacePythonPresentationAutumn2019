import json
import openpyxl
from collections import defaultdict
from IPspace_J2template import create_port_cfg

tree = lambda: defaultdict(tree)

def info_from_xls():

    mig_info = tree()
    wb = openpyxl.load_workbook('Migration_IPspace.xlsx')
    sheet = wb['Vlaninfo']

    finished = False
    i = 1
    while not finished:
        
        hostname = sheet.cell(row=i+1, column=1).value
        vlan = str(sheet.cell(row=i+1, column=2).value) # convert excel to str
       
        # Read interface specific info from excel
        int_descr = sheet.cell(row=i+1, column=4).value
        ip_address_plus_mask = sheet.cell(row=i+1, column=5).value
        ip_helper = sheet.cell(row=i+1, column=6).value
        ip_vrf = sheet.cell(row=i+1, column=7).value
        hsrp_grp = sheet.cell(row=i+1, column=8).value.split(',')[0].split()[0]
        hsrp_gw = sheet.cell(row=i+1, column=8).value.split(',')[0].split()[2]

        # Preprocessing data
        ip_address_plus_mask = ip_address_plus_mask.split()
        ip_address = ip_address_plus_mask[0]
        ip_address_mask = ip_address_plus_mask[1]
                
        # Store migration info in dict
        mig_info[hostname][vlan]['int_descr'] = int_descr
        mig_info[hostname][vlan]['ip_address_mask'] = ip_address_mask
        mig_info[hostname][vlan]['ip_vrf'] = ip_vrf
        mig_info[hostname][vlan]['hsrp_gw'] = hsrp_gw
        mig_info[hostname][vlan]['hsrp_grp'] = hsrp_grp
        
        # Calculate new IP addresses and add to dict
        ip_add = ip_address.split('.')
        ip_add_new_2 = str(int(ip_add[1]) + 107)
        ip_add_new_3 = str(int(ip_add[2]) + int(hsrp_grp))
        ip_add_new = [ip_add[0], ip_add_new_2, ip_add_new_3, ip_add[3]]
        ip_add_new = '.'.join(ip_add_new)
        mig_info[hostname][vlan]['ip_add'] = ip_add_new

        hsrp_gw = hsrp_gw.split('.')
        hsrp_gw_new = [hsrp_gw[0], ip_add_new_2, ip_add_new_3, hsrp_gw[3]]
        hsrp_gw_new = '.'.join(hsrp_gw_new)
        mig_info[hostname][vlan]['hsrp_gw'] = hsrp_gw_new
        mig_info[hostname][vlan]['hsrp_grp'] = ip_add_new_3

        if ip_helper is not None:
            ip_helpers = ['1.2.3.4', '1.2.3.5']
            mig_info[hostname][vlan]['ip_helpers'] = ip_helpers
        
        i += 1
        if sheet.cell(row=i+1, column=1).value == None:
            finished = True

    # Filter out info and change hostnames per requirements
    vlans_ignore = ['23', '24']

    # Make list out of dict to enable changing dict during iteration
    for hostname in list(mig_info):
        # Make list out of dict to enable changing dict during iteration again
        # Filter VLAN's and VRF's
        for vlan in list(mig_info[hostname]):
            # Just in case testing (None key).
            if (mig_info[hostname][vlan].get('ip_vrf') != 'ipspace'
                    and mig_info[hostname][vlan].get('ip_vrf') is not None):
                 mig_info[hostname].pop(vlan)
            # Study this comprehension yourself
            if any(vlan for vlan_ign in vlans_ignore if vlan in vlan_ign):
                mig_info[hostname].pop(vlan)

        # Change hostnames
        if hostname == 'RTR-a':
            mig_info['RTR-c'] = mig_info.pop('RTR-a')
        elif hostname == 'RTR-b':
            mig_info['RTR-d'] = mig_info.pop('RTR-b')
            
    print(json.dumps(mig_info, indent=4))
    return mig_info


def create_mig_files(mig_info):

    with open('RTR-c-cfg.txt', 'w') as rtr_c , \
         open('RTR-d-cfg.txt', 'w') as rtr_d:

        # You could call a function here which generates
        # Global configuration parts
        print('Configuration for RTR-c:', file=rtr_c)
        print(file=rtr_c)
        print('Configuration for RTR-d:', file=rtr_d)
        print(file=rtr_d)

        for hostname in mig_info:
             for vlan in mig_info[hostname]:
                                            
                # Get info from migration dict
                int_descr = mig_info[hostname][vlan]['int_descr']
                ip_add = mig_info[hostname][vlan]['ip_add']
                ip_address_mask = mig_info[hostname][vlan]['ip_address_mask']
                hsrp_gw = mig_info[hostname][vlan]['hsrp_gw'] 
                hsrp_grp = mig_info[hostname][vlan]['hsrp_grp']
                ip_vrf = mig_info[hostname][vlan]['ip_vrf']
                ip_helpers = mig_info[hostname][vlan].get('ip_helpers')

                if hostname == 'RTR-c':
                    
                    # Call function to create interface config for RTR-c         
                    intf_config_rtr_c = create_port_cfg(
                        'Vlan'+vlan,
                        description=int_descr,
                        ipv4_address=ip_add,
                        ip_vrf=ip_vrf,
                        ipv4_address_mask=ip_address_mask,
                        ip_helpers=ip_helpers,
                        hsrp_gateway_ip=hsrp_gw,
                        hsrp_grp_id=hsrp_grp
                    )
                    print(intf_config_rtr_c, file=rtr_c)

                elif hostname == 'RTR-d':

                # Call function to create interface config for RTR-d   
                    intf_config_rtr_d = create_port_cfg(
                        'Vlan'+vlan,
                        description=int_descr,
                        ipv4_address=ip_add,
                        ip_vrf=ip_vrf,
                        ipv4_address_mask=ip_address_mask,
                        ip_helpers=ip_helpers,
                        hsrp_active_router=False,
                        hsrp_gateway_ip=hsrp_gw,
                        hsrp_grp_id=hsrp_grp
                    )

                    print(intf_config_rtr_d, file=rtr_d)


if __name__ == "__main__":


    mig_info = info_from_xls()

    create_mig_files(mig_info)





